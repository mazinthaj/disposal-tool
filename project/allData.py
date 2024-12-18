import pandas as pd
import openpyxl
from pyzbar import pyzbar
import tkinter as tk
from tkinter import filedialog, messagebox
import cv2
import os
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Border, Side, Alignment, Protection
from PIL import Image, ImageTk
from PIL.ExifTags import TAGS
from datetime import datetime

"""
To install the necessary packages, run:
pip install pandas
pip install tk
pip install pyzbar
pip install opencv-python
pip install openpyxl
pip install Pillow
"""

class BarcodeApp:
    def __init__(self):
        self.excel_path = ""
        self.image_directory = ""
        
        self.window = tk.Tk()
        self.window.title("Disposal Tool")
        self.window.geometry('500x450')
        self.window.configure(bg="#ffffff")

        self.header_frame = tk.Frame(self.window, bg="#ffcc00", height=100)
        self.header_frame.pack(fill="x")
        
        self.header_image = Image.open("project/dhl_log.png")
        self.header_image = self.header_image.resize((152, 26), Image.Resampling.LANCZOS)
        self.header_image = ImageTk.PhotoImage(self.header_image)
        
        self.header_image_label = tk.Label(self.header_frame, image=self.header_image, bg="#ffcc00")
        self.header_image_label.pack(pady=10)

        self.title_label = tk.Label(self.window, text="Disposal Tool", font=("Arial", 16, "bold"), fg="#666666", bg="#ffffff")
        self.title_label.pack(pady=10)

        background_frame = tk.Frame(self.window, bg="#ffffff")
        background_frame.pack(fill="both", expand=True, padx=20, pady=20)

        self.excel_file_path_label = tk.Label(background_frame, bg="#ffffff", text="Excel File Path:", pady=10, font=("Arial", 10, "bold"), fg="#666666")
        self.excel_file_path_label.pack()

        self.browse_excel_button = tk.Button(background_frame, bg="#d40511", fg="#ffffff", text="Browse Excel File", command=self.chooseExcelFile, relief="flat", borderwidth=0, font=("Arial", 10, "bold"), padx=10, pady=5, cursor="hand2", width=20)
        self.browse_excel_button.pack(pady=5)
        self.add_button_hover_effect(self.browse_excel_button)

        self.image_folder_path_label = tk.Label(background_frame, bg="#ffffff", text="Image folder Path:", pady=10, font=("Arial", 10, "bold"), fg="#666666")
        self.image_folder_path_label.pack()

        self.browse_image_button = tk.Button(background_frame, bg="#d40511", fg="#ffffff", text="Browse Image Folder", command=self.browse_directory, relief="flat", borderwidth=0, font=("Arial", 10, "bold"), padx=10, pady=5, cursor="hand2", width=20)
        self.browse_image_button.pack(pady=5)
        self.add_button_hover_effect(self.browse_image_button)

        self.process_button = tk.Button(background_frame, bg="#d40511", fg="#ffffff", text="Process Files", font=("Arial", 12, "bold"), padx=10, pady=10, command=self.process_files, relief="flat", borderwidth=0, cursor="hand2", width=20)
        self.process_button.pack(pady=30)
        self.add_button_hover_effect(self.process_button)

    def add_button_hover_effect(self, button):
        """ Adds hover effect to buttons. """
        def on_enter(event):
            button.config(bg="#b2050f")

        def on_leave(event):
            button.config(bg="#d40511")

        button.bind("<Enter>", on_enter)
        button.bind("<Leave>", on_leave)

    def scan_barcode(self, image):
        """ Scans all barcodes in the image, but only returns the CODE39 barcode. """
        barcodes = pyzbar.decode(image)
        for barcode in barcodes:
            barcode_data = barcode.data.decode("utf-8")
            if barcode.type == "CODE39":
                return True, barcode_data
        return False, ""

    def get_image_creation_date(self, image_path):
        """ Gets the image creation date. If no EXIF data, returns the file creation date. """
        try:
            with Image.open(image_path) as img:
                exif = img._getexif()
                if exif:
                    exif = {TAGS.get(key, key): value for key, value in exif.items()}
                    for field in ['DateTimeOriginal', 'DateTimeDigitized', 'DateTime']:
                        if field in exif:
                            try:
                                return datetime.strptime(exif[field], '%Y:%m:%d %H:%M:%S')
                            except:
                                continue
            stats = os.stat(image_path)
            return datetime.fromtimestamp(stats.st_ctime)
        except:
            return datetime.max

    def chooseExcelFile(self):
        """ Allows user to select only Excel file. """
        self.excel_path = filedialog.askopenfilename(
            title="Select a file",
            filetypes=[("Excel files", "*.xlsx")]
        )
        self.excel_file_path_label.config(text=f"Excel File Path: {self.excel_path}", pady=10, font=("Arial", 10, "bold"), fg="#ffae00")

    def browse_directory(self):
        """ Allows user to choose the image directory. """
        self.image_directory = filedialog.askdirectory()
        if self.image_directory:
            self.image_folder_path_label.config(text=f"Image folder Path: {self.image_directory}", pady=10, font=("Arial", 10, "bold"), fg="#ffae00")

    def find_header_row(self, worksheet):
        """ Finds the row index that contains 'HWB/SID' in the first column. """
        for row in range(1, worksheet.max_row + 1):
            if worksheet.cell(row=row, column=1).value == 'HWB/SID':
                return row
        return None

    def process_sheet(self, wb, sheet_name, data):
        """ Applies styles to the sheet and updates it with image links. """
        ws = wb[sheet_name]
        header_row = self.find_header_row(ws)
        if header_row is None:
            tk.messagebox.showwarning("Warning", f"!!!! Could not find 'HWB/SID' column in sheet '{sheet_name}'")
            return False

        original_formatting = {}
        for row in ws.iter_rows(min_row=header_row):
            for cell in row:
                if cell.has_style:
                    original_formatting[(cell.row, cell.column)] = {
                        'font': Font(
                            name=cell.font.name,
                            size=cell.font.size,
                            bold=cell.font.bold,
                            italic=cell.font.italic,
                            vertAlign=cell.font.vertAlign,
                            underline=cell.font.underline,
                            strike=cell.font.strike,
                            color=cell.font.color
                        ),
                        'fill': PatternFill(
                            fill_type=cell.fill.fill_type,
                            start_color=cell.fill.start_color,
                            end_color=cell.fill.end_color
                        ),
                        'border': Border(
                            left=Side(border_style=cell.border.left.style, color=cell.border.left.color),
                            right=Side(border_style=cell.border.right.style, color=cell.border.right.color),
                            top=Side(border_style=cell.border.top.style, color=cell.border.top.color),
                            bottom=Side(border_style=cell.border.bottom.style, color=cell.border.bottom.color)
                        ),
                        'alignment': Alignment(
                            horizontal=cell.alignment.horizontal,
                            vertical=cell.alignment.vertical,
                            text_rotation=cell.alignment.text_rotation,
                            wrap_text=cell.alignment.wrap_text,
                            shrink_to_fit=cell.alignment.shrink_to_fit,
                            indent=cell.alignment.indent
                        ),
                        'protection': Protection(
                            locked=cell.protection.locked,
                            hidden=cell.protection.hidden
                        ),
                        'number_format': cell.number_format
                    }

        df = pd.read_excel(self.excel_path, sheet_name=sheet_name, header=header_row-1)
        df.dropna(how='all', inplace=True)
        df.drop_duplicates(inplace=True)
        df = df.dropna(subset=['HWB/SID'])

        if 'Before' not in df.columns:
            df['Before'] = None
        if 'After' not in df.columns:
            df['After'] = None

        for index, row in df.iterrows():
            sid_number = str(int(row['HWB/SID']))
            if sid_number in data:
                df.loc[index, 'Before'] = f'=HYPERLINK("{data[sid_number][0]}", "Click to view image")'
                if len(data[sid_number]) > 1:
                    df.loc[index, 'After'] = f'=HYPERLINK("{data[sid_number][-1]}", "Click to view image")'
                else:
                    df.loc[index, 'After'] = "IMAGE NOT FOUND"
            else:
                df.loc[index, 'Before'] = "IMAGE NOT FOUND"
                df.loc[index, 'After'] = "IMAGE NOT FOUND"

        for col_idx, column_name in enumerate(df.columns, 1):
            cell = ws.cell(row=header_row, column=col_idx)
            if (header_row, col_idx) in original_formatting:
                orig_format = original_formatting[(header_row, col_idx)]
                cell.font = orig_format['font']
                cell.fill = orig_format['fill']
                cell.border = orig_format['border']
                cell.alignment = orig_format['alignment']
                cell.protection = orig_format['protection']
                cell.number_format = orig_format['number_format']
            cell.value = column_name

        num = 0
        for idx, row in df.iterrows():
            current_row = header_row + 1 + idx
            for col_idx, value in enumerate(row, 1):
                cell = ws.cell(row=current_row, column=col_idx)
                if (current_row, col_idx) in original_formatting:
                    orig_format = original_formatting[(current_row, col_idx)]
                    cell.font = orig_format['font']
                    cell.fill = orig_format['fill']
                    cell.border = orig_format['border']
                    cell.alignment = orig_format['alignment']
                    cell.protection = orig_format['protection']
                    cell.number_format = orig_format['number_format']
                else:
                    if num < 2:
                        ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = 20
                        num += 1
                cell.value = value
                if ('Before' in df.columns and col_idx == df.columns.get_loc('Before') + 1) or \
                ('After' in df.columns and col_idx == df.columns.get_loc('After') + 1):
                    if 'HYPERLINK' in value:
                        cell.font = Font(color="0000FF", underline="single")
                    elif value == "IMAGE NOT FOUND":
                        cell.font = Font(color="FF0000", bold=True)

        return True

    def process_files(self):
        """ 
        Creates a dict data. Stores the barcode as a key and their locations in a list. Renames the files by sorting according  
        to their creation date. Chooses the first image as before and the last image as the after.
        Checks if the barcode is there in the excel and displays the image location.
        """
        data = {}
        if self.excel_path and self.image_directory:
            image_files = []
            for file in os.listdir(self.image_directory):
                if file.lower().endswith((".jpg", ".png", ".jpeg")):
                    full_path = os.path.join(self.image_directory, file)
                    creation_date = self.get_image_creation_date(full_path)
                    image_files.append((file, creation_date))
            
            if not image_files:
                tk.messagebox.showwarning("Warning", "!!!! No images of type .jpg, .png, .jpeg found")
                return

            image_files.sort(key=lambda x: x[1])
            sorted_names = [name for name, _ in image_files]

            for file_name in sorted_names:
                image_path = os.path.join(self.image_directory, file_name)
                image = cv2.imread(image_path)
                blurred = cv2.GaussianBlur(image, (5, 5), 0)
                gray = cv2.cvtColor(blurred, cv2.COLOR_BGR2GRAY)
                flag, image_barcode = self.scan_barcode(gray)
                
                if flag:
                    if image_barcode not in data:
                        new_name = os.path.join(self.image_directory, f"{image_barcode}_before{os.path.splitext(file_name)[1]}")
                        os.rename(image_path, new_name)
                        data[image_barcode] = [new_name]
                    else:
                        new_name = os.path.join(self.image_directory, f"{image_barcode}({len(data[image_barcode])})_after{os.path.splitext(file_name)[1]}")
                        os.rename(image_path, new_name)
                        data[image_barcode].append(new_name)

            wb = load_workbook(self.excel_path)
            processed_sheets = 0
            
            for sheet_name in wb.sheetnames:
                if self.process_sheet(wb, sheet_name, data):
                    processed_sheets += 1

            if processed_sheets > 0:
                wb.save(self.excel_path)
                tk.messagebox.showinfo("Success", f"Processing completed successfully! Processed {processed_sheets} sheets.")
            else:
                tk.messagebox.showwarning("Warning", "!!!! No sheets were processed successfully")
        else:
            tk.messagebox.showwarning("Warning", "!!!! Please select both Excel file and image directory first")

    def run(self):
        self.window.mainloop()

# Main code
if __name__ == "__main__":
    app = BarcodeApp()
    app.run()
