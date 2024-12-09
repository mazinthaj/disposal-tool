import pandas as pd
import openpyxl
from pyzbar import pyzbar
import tkinter as tk
from tkinter import filedialog, messagebox
import cv2
import os
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Border, Side, Alignment, Protection
from PIL import Image, ImageTk
from pillow_heif import register_heif_opener

""" 
pip install pandas
pip install tk
pip install pyzbar
pip install opencv-python
pip install openpyxl
pip install pillow
pip3 install pillow-heif
"""

""" 
Warnings:
When taking the picture of the barcode, the waybil barcode must be shown completely
before and after image file should exist.
The after image must be taken after the before image. (If 2 barcode photos are taken one after the other the 2nd one becomes the first one's after photo)
take pictures in heic file format (IPhone).
do not delete the jpg file created by the program.
"""

class BarcodeApp:
    def __init__(self):
        self.excel_path = ""
        self.image_directory = ""
        
        self.window = tk.Tk()
        self.window.title("Disposal Tool")
        self.window.geometry('500x450')
        self.window.configure(bg="#ffffff")

        self.header_frame = tk.Frame(self.window, bg="#ffcc00", height=100)
        self.header_frame.pack(fill="x")
        
        self.header_image = Image.open(r"C:\Users\mthajude\Downloads\project\dhl_log.png")
        self.header_image = self.header_image.resize((152, 26), Image.Resampling.LANCZOS)
        self.header_image = ImageTk.PhotoImage(self.header_image)
        
        self.header_image_label = tk.Label(self.header_frame, image=self.header_image, bg="#ffcc00")
        self.header_image_label.pack(pady=10)

        self.title_label = tk.Label(self.window, text="Disposal Tool", font=("Arial", 16, "bold"), fg="#666666", bg="#ffffff")
        self.title_label.pack(pady=10)

        background_frame = tk.Frame(self.window, bg="#ffffff")
        background_frame.pack(fill="both", expand=True, padx=20, pady=20)

        self.excel_file_path_label = tk.Label(background_frame, bg="#ffffff", text="Excel File Path:", pady=10, font=("Arial", 10, "bold"), fg="#666666")
        self.excel_file_path_label.pack()

        self.browse_excel_button = tk.Button(background_frame, bg="#d40511", fg="#ffffff", text="Browse Excel File", command=self.chooseExcelFile, relief="flat", borderwidth=0, font=("Arial", 10, "bold"), padx=10, pady=5, cursor="hand2", width=20)
        self.browse_excel_button.pack(pady=5)
        self.add_button_hover_effect(self.browse_excel_button)

        self.image_folder_path_label = tk.Label(background_frame, bg="#ffffff", text="Image folder Path:", pady=10, font=("Arial", 10, "bold"), fg="#666666")
        self.image_folder_path_label.pack()

        self.browse_image_button = tk.Button(background_frame, bg="#d40511", fg="#ffffff", text="Browse Image Folder", command=self.browse_directory, relief="flat", borderwidth=0, font=("Arial", 10, "bold"), padx=10, pady=5, cursor="hand2", width=20)
        self.browse_image_button.pack(pady=5)
        self.add_button_hover_effect(self.browse_image_button)

        self.process_button = tk.Button(background_frame, bg="#d40511", fg="#ffffff", text="Process Files", font=("Arial", 12, "bold"), padx=10, pady=10, command=self.process_files, relief="flat", borderwidth=0, cursor="hand2", width=20)
        self.process_button.pack(pady=30)
        self.add_button_hover_effect(self.process_button)

    def add_button_hover_effect(self, button):
        """ Adds hover effect to buttons. """
        def on_enter(event):
            button.config(bg="#b2050f")

        def on_leave(event):
            button.config(bg="#d40511")

        button.bind("<Enter>", on_enter)
        button.bind("<Leave>", on_leave)

        
    #Scans all barcodes in the image, but only sends the CODE39 barcode
    def scan_barcode(self, image):
        barcodes = pyzbar.decode(image)

        for barcode in barcodes:
            barcode_data = barcode.data.decode("utf-8")
            barcode_type = barcode.type

            if barcode_type == "CODE39": 
               return True, barcode_data
        return False, ""

    #allows user to select only excel file
    def chooseExcelFile(self):
        self.excel_path = filedialog.askopenfilename(
            title="Select a file",
            filetypes=[
                ("Excel files", "*.xlsx")
            ]
        )
        self.excel_file_path_label.config(text=f"Excel File Path:   {self.excel_path}", pady=10, font=("Arial", 10, "bold"), fg="#ffae00")

    #allows user to choose the image directory
    def browse_directory(self):
        self.image_directory = filedialog.askdirectory()
        if self.image_directory:
            self.image_folder_path_label.config(text=f"Image folder Path:   {self.image_directory}", pady=10, font=("Arial", 10, "bold"), fg="#ffae00")

    #Finds the row index that contains 'HWB/SID' in the first column as this is the most important row that displays data
    def find_header_row(self, worksheet):
        for row in range(1, worksheet.max_row + 1):
            if worksheet.cell(row=row, column=1).value == 'HWB/SID':
                return row
        return None

    #applying styles to the sheet
    def process_sheet(self, wb, sheet_name, data):
        ws = wb[sheet_name]
        header_row = self.find_header_row(ws)

        if header_row is None:
            tk.messagebox.showwarning("Warning", f"!!!! Could not find 'HWB/SID' column in sheet '{sheet_name}'")
            return False

        # Store original cell formatting before modifying the worksheet
        original_formatting = {}
        for row in ws.iter_rows(min_row=header_row):
            for cell in row:
                if cell.has_style:
                    original_formatting[(cell.row, cell.column)] = {
                        'font': Font(
                            name=cell.font.name,
                            size=cell.font.size,
                            bold=cell.font.bold,
                            italic=cell.font.italic,
                            vertAlign=cell.font.vertAlign,
                            underline=cell.font.underline,
                            strike=cell.font.strike,
                            color=cell.font.color
                        ),
                        'fill': PatternFill(
                            fill_type=cell.fill.fill_type,
                            start_color=cell.fill.start_color,
                            end_color=cell.fill.end_color
                        ),
                        'border': Border(
                            left=Side(border_style=cell.border.left.style, color=cell.border.left.color),
                            right=Side(border_style=cell.border.right.style, color=cell.border.right.color),
                            top=Side(border_style=cell.border.top.style, color=cell.border.top.color),
                            bottom=Side(border_style=cell.border.bottom.style, color=cell.border.bottom.color)
                        ),
                        'alignment': Alignment(
                            horizontal=cell.alignment.horizontal,
                            vertical=cell.alignment.vertical,
                            text_rotation=cell.alignment.text_rotation,
                            wrap_text=cell.alignment.wrap_text,
                            shrink_to_fit=cell.alignment.shrink_to_fit,
                            indent=cell.alignment.indent
                        ),
                        'protection': Protection(
                            locked=cell.protection.locked,
                            hidden=cell.protection.hidden
                        ),
                        'number_format': cell.number_format
                    }

        # Read Excel file with the correct header row for this sheet
        df = pd.read_excel(self.excel_path, sheet_name=sheet_name, header=header_row-1)

        df.dropna(how='all', inplace=True)
        df.drop_duplicates(inplace=True)
        df = df.dropna(subset=['HWB/SID'])

        if 'Before' not in df.columns:
            df['Before'] = None
        if 'After' not in df.columns:
            df['After'] = None

        # Update DataFrame with image links
        for index, row in df.iterrows():
            sid_number = str(int((row['HWB/SID'])))
            if sid_number in data:
                df.loc[index, 'Before'] = f'=HYPERLINK("{data[sid_number][0]}", "Click to view image")'
                df.loc[index, 'After'] = f'=HYPERLINK("{data[sid_number][1]}", "Click to view image")'
            else:
                df.loc[index, 'Before'] = "IMAGE NOT FOUND"
                df.loc[index, 'After'] = "IMAGE NOT FOUND"

        #Update styling for header rows
        for col_idx, column_name in enumerate(df.columns, 1):
            cell = ws.cell(row=header_row, column=col_idx)
            
            if (header_row, col_idx) in original_formatting:
                orig_format = original_formatting[(header_row, col_idx)]
                cell.font = orig_format['font']
                cell.fill = orig_format['fill']
                cell.border = orig_format['border']
                cell.alignment = orig_format['alignment']
                cell.protection = orig_format['protection']
                cell.number_format = orig_format['number_format']
            
            cell.value = column_name

        #Update styling for data rows
        num = 0
        for idx, row in df.iterrows():
            current_row = header_row + 1 + idx
            for col_idx, value in enumerate(row, 1):
                cell = ws.cell(row=current_row, column=col_idx)
                
                # Preserve original formatting if it existed
                if (current_row, col_idx) in original_formatting:
                    orig_format = original_formatting[(current_row, col_idx)]
                    cell.font = orig_format['font']
                    cell.fill = orig_format['fill']
                    cell.border = orig_format['border']
                    cell.alignment = orig_format['alignment']
                    cell.protection = orig_format['protection']
                    cell.number_format = orig_format['number_format']
                else:
                    if num <2:
                        ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = 20
                        num+=1
                cell.value = value
                
                # Apply special formatting for hyperlinks and "IMAGE NOT FOUND"
                if ('Before' in df.columns and col_idx == df.columns.get_loc('Before') + 1) or \
                ('After' in df.columns and col_idx == df.columns.get_loc('After') + 1):
                    if 'HYPERLINK' in value:
                        cell.font = Font(color="0000FF", underline="single")
                    elif value == "IMAGE NOT FOUND":
                        cell.font = Font(color="FF0000", bold=True)

        return True
    
    """ 
    Creates a dict data. Stores the barcode as a key and their locations in a list. Renames the files. 
    Chooses the first image as before and the second image as the after.
    Checks if the barcode is there in the excel and displays the image location.
    """
    def process_files(self):
        data = {}
        if self.excel_path and self.image_directory:
            heic_files = [f for f in os.listdir(self.image_directory) if f.endswith(('.heic', '.jpg'))]
            if not heic_files:
                tk.messagebox.showwarning("Warning", "!!!! No images of type .heic, .jpg found")
                return
            register_heif_opener()
            processed_files = []

            for f in heic_files:
                full_path = os.path.join(self.image_directory, f)

                base_path = os.path.splitext(full_path)[0]
                jpg_path = base_path + '.jpg'

                if os.path.exists(jpg_path):
                    processed_files.append(jpg_path)
                    continue

                image = Image.open(full_path)
                image.save(jpg_path, "JPEG")
                processed_files.append(jpg_path)

            processed_files.sort()

            i = 0
            while i < len(processed_files):
                image = cv2.imread(processed_files[i])
                blurred = cv2.GaussianBlur(image, (5, 5), 0)
                gray = cv2.cvtColor(blurred, cv2.COLOR_BGR2GRAY)

                flag, image_barcode = self.scan_barcode(gray)

                if flag:
                    new_before = os.path.join(self.image_directory, f"{image_barcode}_before.jpg")
                    os.replace(processed_files[i], new_before)

                    if i + 1 < len(processed_files):
                        new_after = os.path.join(self.image_directory, f"{image_barcode}_after.jpg")
                        os.replace(processed_files[i + 1], new_after)
                        data[image_barcode] = [new_before, new_after]
                        i += 1
                i += 1

            # Process each sheet in the workbook
            wb = load_workbook(self.excel_path)
            processed_sheets = 0

            for sheet_name in wb.sheetnames:
                if self.process_sheet(wb, sheet_name, data):
                    processed_sheets += 1

            if processed_sheets > 0:
                wb.save(self.excel_path)
                tk.messagebox.showinfo("Success", f"Processing completed successfully! Processed {processed_sheets} sheets.")
            else:
                tk.messagebox.showwarning("Warning", "!!!! No sheets were processed successfully")
        else:
            tk.messagebox.showwarning("Warning", "!!!! Please select both Excel file and image directory first")


    def run(self):
        self.window.mainloop()

#main code ..................................................................................
app = BarcodeApp()
app.run()
